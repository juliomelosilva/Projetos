import openpyxl
import time
import pyautogui
import webbrowser
import keyboard
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import json
import os
import threading

# ==================================================
# CONFIGURAÇÕES GERAIS
# ==================================================

SITE_URL = "https://smliveloja.bitrix24.site/plantao/"
CONFIG_DIA = "config_dia.json"

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3

# ==================================================
# VARIÁVEIS GLOBAIS
# ==================================================

IMAGENS = {}
PLANILHA = ""

# ==================================================
# FUNÇÃO DE CLIQUE POR IMAGEM
# ==================================================

def clicar_imagem(nome, timeout=15, confidence=0.8):
    inicio = time.time()
    while time.time() - inicio < timeout:
        pos = pyautogui.locateCenterOnScreen(
            IMAGENS[nome],
            confidence=confidence
        )
        if pos:
            pyautogui.click(pos)
            return
        time.sleep(0.4)
    raise Exception(f"Imagem não encontrada: {nome}")

# ==================================================
# LIMPAR PLANILHA (APENAS DA LINHA 2 PARA BAIXO)
# ==================================================

def limpar_planilha(sheet, wb):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    wb.save(PLANILHA)

# ==================================================
# DESLIGAMENTO COM CONFIRMAÇÃO + AUTO 10s
# ==================================================

def desligar_com_confirmacao():

    desligar_confirmado = {"cancelado": False, "respondido": False}

    def perguntar():
        resposta = messagebox.askokcancel(
            "Desligamento",
            "Todos os chamados foram registrados com sucesso.\n\n"
            "Desligamento solicitado, deseja cancelar?"
        )

        desligar_confirmado["respondido"] = True

        if resposta:
            os.system("shutdown /s /t 0")
        else:
            desligar_confirmado["cancelado"] = True
            countdown_label.config(text="Desligamento cancelado.")

    def contagem():
        for i in range(10, 0, -1):
            if desligar_confirmado["respondido"]:
                return
            countdown_label.config(text=f"Desligando automaticamente em {i}s...")
            time.sleep(1)

        if not desligar_confirmado["cancelado"]:
            os.system("shutdown /s /t 0")

    threading.Thread(target=perguntar).start()
    threading.Thread(target=contagem).start()

# ==================================================
# SELEÇÃO DO DIA
# ==================================================

def salvar_dia(coord):
    with open(CONFIG_DIA, "w") as f:
        json.dump(coord, f)

def carregar_dia():
    if os.path.exists(CONFIG_DIA):
        with open(CONFIG_DIA, "r") as f:
            return json.load(f)
    return None

def configurar_dia():
    clicar_imagem("selecao_data.png")

    messagebox.showinfo(
        "Configurar Dia",
        "Posicione o mouse sobre o DIA desejado\n"
        "e pressione F8 para gravar"
    )

    keyboard.wait("f8")

    x, y = pyautogui.position()
    salvar_dia({"x": x, "y": y})

    messagebox.showinfo(
        "Sucesso",
        f"Dia gravado em:\nX={x} | Y={y}"
    )

# ==================================================
# INTERFACE – SELEÇÃO DE ARQUIVOS
# ==================================================

def selecionar_imagens():
    arquivos = filedialog.askopenfilenames(
        title="Selecione todas as imagens",
        filetypes=[("PNG", "*.png")]
    )
    for img in arquivos:
        IMAGENS[os.path.basename(img)] = img

    messagebox.showinfo(
        "Imagens carregadas",
        f"{len(IMAGENS)} imagens carregadas"
    )

def selecionar_planilha():
    global PLANILHA
    PLANILHA = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Excel", "*.xlsx")]
    )

# ==================================================
# AUTOMAÇÃO PRINCIPAL
# ==================================================

def iniciar_automacao():
    nome_tecnico = nome_entry.get()

    if not nome_tecnico:
        messagebox.showerror("Erro", "Informe o nome do técnico")
        return

    if not PLANILHA:
        messagebox.showerror("Erro", "Selecione a planilha")
        return

    dia = carregar_dia()
    if not dia:
        messagebox.showerror("Erro", "Configure o dia antes de iniciar")
        return

    wb = openpyxl.load_workbook(PLANILHA)
    sheet = wb.active

    total = sum(1 for row in sheet.iter_rows(min_row=2) if row[0].value)
    progress_bar["maximum"] = total
    progress_bar["value"] = 0

    contador = 0

    for linha in sheet.iter_rows(min_row=2):

        empresa = linha[0].value
        titulo = linha[1].value
        descricao = linha[1].value
        resolucao = linha[2].value

        if not empresa:
            break

        clicar_imagem("tecnico.png")
        keyboard.write(str(nome_tecnico))
        time.sleep(0.4)

        clicar_imagem("selecao_data.png")
        time.sleep(0.4)
        pyautogui.click(dia["x"], dia["y"])
        time.sleep(0.4)

        clicar_imagem("empresa.png")
        keyboard.write(str(empresa))
        time.sleep(0.3)

        clicar_imagem("titulo.png")
        keyboard.write(str(titulo))
        time.sleep(0.3)

        clicar_imagem("descricao.png")
        keyboard.write(str(descricao))
        time.sleep(0.3)

        clicar_imagem("resolucao.png")
        keyboard.write(str(resolucao))
        time.sleep(0.3)

        pyautogui.scroll(-100000)
        time.sleep(0.3)

        clicar_imagem("prioridade.png")
        time.sleep(0.4)

        clicar_imagem("prioridade_opcao.png")
        time.sleep(0.4)

        clicar_imagem("categoria.png")
        time.sleep(0.6)

        pyautogui.moveRel(0, 100)
        time.sleep(0.3)

        for _ in range(5):
            keyboard.press_and_release('end')
            time.sleep(0.2)

        for _ in range(9):
            pyautogui.scroll(-100000)
            time.sleep(0.15)

        pyautogui.moveRel(500, 0)
        pyautogui.scroll(-100000)
        time.sleep(0.15)

        clicar_imagem("categoria_opcao.png")
        time.sleep(0.3)

        clicar_imagem("finalizar.png")
        time.sleep(6)

        contador += 1
        progress_bar["value"] = contador
        app.update()

    limpar_planilha(sheet, wb)

    if desligar_var.get():
        desligar_com_confirmacao()

# ==================================================
# INTERFACE
# ==================================================

app = tk.Tk()
app.title("Bot Chamados Plantão")
app.geometry("360x360")

webbrowser.open(SITE_URL)

tk.Label(app, text="Nome do Técnico").pack(pady=5)
nome_entry = tk.Entry(app, width=30)
nome_entry.pack()

tk.Button(app, text="Selecionar Imagens", command=selecionar_imagens).pack(pady=5)
tk.Button(app, text="Selecionar Planilha", command=selecionar_planilha, bg="#0d3303", fg="white").pack(pady=5)
tk.Button(app, text="Configurar Dia (F8)", command=configurar_dia, bg="#ffaa00").pack(pady=10)

desligar_var = tk.BooleanVar()
tk.Checkbutton(
    app,
    text="Desligar computador ao finalizar",
    variable=desligar_var
).pack(pady=5)

progress_bar = ttk.Progressbar(app, length=260)
progress_bar.pack(pady=10)

countdown_label = tk.Label(app, text="", fg="red")
countdown_label.pack()

tk.Button(app, text="Iniciar Automação", command=iniciar_automacao, bg="green", fg="white").pack(pady=10)

app.mainloop()
