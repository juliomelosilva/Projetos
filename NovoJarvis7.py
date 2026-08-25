import openpyxl
import time
import pyautogui
import webbrowser
import keyboard
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import json
import os
import threading
import sys

# ==================================================
# CONFIGURAÇÕES GERAIS
# ==================================================

SITE_URL = "https://api.leadconnectorhq.com/widget/form/UGgmkEe3YBsaDMsI58pG"
CONFIG_DIA = "config_dia.json"

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3

# ==================================================
# VARIÁVEIS GLOBAIS
# ==================================================

IMAGENS = {}
PLANILHA = ""
JANELA_ABERTA = True

# ==================================================
# FUNÇÕES DE INTERAÇÃO NA TELA
# ==================================================

def clicar_imagem(nome, timeout=15, confidence=0.8):
    inicio = time.time()
    while time.time() - inicio < timeout:
        if not JANELA_ABERTA:
            return
        try:
            pos = pyautogui.locateCenterOnScreen(
                IMAGENS[nome],
                confidence=confidence
            )
            if pos:
                pyautogui.click(pos)
                return
        except Exception:
            pass
        time.sleep(0.4)
    raise Exception(f"Imagem não encontrada: {nome}")

def tentar_clicar_recomecar():
    if "recomecar.png" not in IMAGENS:
        return

    try:
        pos = pyautogui.locateCenterOnScreen(
            IMAGENS["recomecar.png"],
            confidence=0.7
        )
        if pos:
            pyautogui.click(pos)
            time.sleep(1.0)
    except Exception:
        pass

def preencher_campo_seguro(imagem_nome, texto, enter_ao_final=False, ajustar_posicao_revenda=False):
    """
    Se ajustar_posicao_revenda=True:
    1. Localiza a imagem e clica.
    2. Obtém a posição atual X, Y e move para Y + 5px (para baixo).
    3. Clica novamente na imagem/campo.
    4. Preenche os dados da planilha.
    """
    if not JANELA_ABERTA or texto is None: 
        return
    
    # 1. Localiza a imagem e clica no campo
    clicar_imagem(imagem_nome)
    time.sleep(0.3)
    
    # 2 e 3. Fluxo específico para a Revenda
    if ajustar_posicao_revenda:
        # =========================================================
        # DESTAQUE: MOVENDO O CURSOR DO MOUSE 5PX PARA BAIXO
        # =========================================================
        x_atual, y_atual = pyautogui.position()
        pyautogui.moveTo(x_atual, y_atual + 100, duration=0.2)
        # =========================================================
        time.sleep(0.3)
        
        # Clica novamente após o movimento
        clicar_imagem(imagem_nome)
        time.sleep(0.3)

    # 4. Limpa e preenche o valor da planilha
    keyboard.press_and_release('ctrl+a')
    keyboard.press_and_release('backspace')
    time.sleep(0.1)
    
    keyboard.write(str(texto))
    time.sleep(0.3)
    
    if enter_ao_final:
        keyboard.press_and_release('enter')
    else:
        keyboard.press_and_release('esc')
    time.sleep(0.2)

# ==================================================
# LIMPAR PLANILHA (APENAS DA LINHA 2 PARA BAIXO)
# ==================================================

def limpar_planilha(sheet, wb):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    wb.save(PLANILHA)

# ==================================================
# DESLIGAMENTO COM CONFIRMAÇÃO + AUTO 10s
# ==================================================

def desligar_com_confirmacao():
    desligar_confirmado = {"cancelado": False, "respondido": False}

    def perguntar():
        resposta = messagebox.askokcancel(
            "Desligamento",
            "Todos os chamados foram registrados com sucesso.\n\n"
            "Desligamento solicitado, deseja cancelar?"
        )
        desligar_confirmado["respondido"] = True
        if resposta:
            os.system("shutdown /s /t 0")
        else:
            desligar_confirmado["cancelado"] = True
            if JANELA_ABERTA:
                countdown_label.config(text="Desligamento cancelado.")

    def contagem():
        for i in range(10, 0, -1):
            if not JANELA_ABERTA or desligar_confirmado["respondido"]:
                return
            countdown_label.config(text=f"Desligando automaticamente em {i}s...")
            time.sleep(1)
        if not desligar_confirmado["cancelado"] and JANELA_ABERTA:
            os.system("shutdown /s /t 0")

    t1 = threading.Thread(target=perguntar, daemon=True)
    t2 = threading.Thread(target=contagem, daemon=True)
    t1.start()
    t2.start()

# ==================================================
# SELEÇÃO DO DIA E CONFIRMAÇÃO
# ==================================================

def salvar_config_data(coord):
    with open(CONFIG_DIA, "w") as f:
        json.dump(coord, f)

def carregar_config_data():
    if os.path.exists(CONFIG_DIA):
        with open(CONFIG_DIA, "r") as f:
            return json.load(f)
    return None

def configurar_dia():
    tentar_clicar_recomecar()
    time.sleep(0.5)

    clicar_imagem("selecao_data.png")
    
    messagebox.showinfo(
        "Configurar Dia - Passo 1",
        "Deixe o mouse sobre o DIA desejado\n"
        "e pressione F8 para gravar."
    )
    keyboard.wait("f8")
    dia_x, dia_y = pyautogui.position()
    time.sleep(0.5)

    messagebox.showinfo(
        "Configurar Dia - Passo 2",
        "Agora deixe o mouse sobre o botão CONFIRMAR\n"
        "e pressione F8 para gravar."
    )
    keyboard.wait("f8")
    conf_x, conf_y = pyautogui.position()

    salvar_config_data({
        "dia": {"x": dia_x, "y": dia_y},
        "confirmar": {"x": conf_x, "y": conf_y}
    })

    pyautogui.click(dia_x, dia_y)
    time.sleep(0.3)
    pyautogui.click(conf_x, conf_y)
    time.sleep(0.5)

    keyboard.press_and_release('f5')
    time.sleep(1.0)

    messagebox.showinfo(
        "Sucesso",
        f"Posições gravadas com sucesso!\n\n"
        f"Dia: X={dia_x} | Y={dia_y}\n"
        f"Confirmar: X={conf_x} | Y={conf_y}\n\n"
        f"A página foi atualizada e está pronta para a automação."
    )

# ==================================================
# INTERFACE – SELEÇÃO DE ARQUIVOS
# ==================================================

def selecionar_imagens():
    arquivos = filedialog.askopenfilenames(
        title="Selecione todas as imagens",
        filetypes=[("PNG", "*.png")]
    )
    for img in arquivos:
        IMAGENS[os.path.basename(img)] = img
    messagebox.showinfo(
        "Imagens carregadas",
        f"{len(IMAGENS)} imagens carregadas"
    )

def selecionar_planilha():
    global PLANILHA
    PLANILHA = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Excel", "*.xlsx")]
    )

# ==================================================
# AUTOMAÇÃO PRINCIPAL
# ==================================================

def iniciar_automacao():
    if not PLANILHA:
        messagebox.showerror("Erro", "Selecione a planilha")
        return

    config_data = carregar_config_data()
    if not config_data or "dia" not in config_data or "confirmar" not in config_data:
        messagebox.showerror("Erro", "Configure a data e o botão confirmar (F8) antes de iniciar")
        return

    tentar_clicar_recomecar()

    wb = openpyxl.load_workbook(PLANILHA)
    sheet = wb.active

    total = sum(1 for row in sheet.iter_rows(min_row=2) if row[0].value)
    progress_bar["maximum"] = total
    progress_bar["value"] = 0
    contador = 0

    for linha in sheet.iter_rows(min_row=2):
        if not JANELA_ABERTA:
            break

        empresa = linha[0].value                                 # Coluna A (Empresa)
        titulo = linha[1].value                                  # Coluna B (Problema / Título)
        descricao = linha[1].value                               # Coluna B (Problema / Descrição)
        resolucao = linha[2].value                               # Coluna C (Resoluçao)
        revenda = linha[3].value if len(linha) > 3 else None     # Coluna D (Revenda)
        segmento = linha[4].value if len(linha) > 4 else None    # Coluna E (Segmento)

        if not empresa:
            break

        # 1. Seleção de Data
        clicar_imagem("selecao_data.png")
        time.sleep(0.3)
        pyautogui.click(config_data["dia"]["x"], config_data["dia"]["y"])
        time.sleep(0.3)
        pyautogui.click(config_data["confirmar"]["x"], config_data["confirmar"]["y"])
        time.sleep(1.0)

        # 2. Empresa
        preencher_campo_seguro("empresa.png", empresa)

        # 3. Título
        preencher_campo_seguro("titulo.png", titulo)

        # 4. Descrição + TAB
        preencher_campo_seguro("descricao.png", descricao)
        keyboard.press_and_release('tab')
        time.sleep(0.2)

        # 5. Resolução
        preencher_campo_seguro("resolucao.png", resolucao)

        # Desce a tela
        pyautogui.scroll(-2000)
        time.sleep(0.5)

        # 6. Segmento
        if segmento and "segmento.png" in IMAGENS:
            preencher_campo_seguro("segmento.png", segmento, enter_ao_final=True)

        # 7. Categoria + TAB
        if "categoria.png" in IMAGENS:
            preencher_campo_seguro("categoria.png", "PLANTÃO", enter_ao_final=True)
            time.sleep(0.2)

        # 8. Revenda (Localiza revenda.png -> Clica -> Move 5px para baixo -> Clica novamente -> Preenche com o dado da planilha)
        if revenda and "revenda.png" in IMAGENS:
            preencher_campo_seguro("revenda.png", revenda, enter_ao_final=True, ajustar_posicao_revenda=True)

        time.sleep(0.5)

        # 9. Finalizar / Enviar Registro
        if "finalizar.png" in IMAGENS:
            clicar_imagem("finalizar.png")
        else:
            clicar_imagem("enviar.png")
        
        time.sleep(4)

        contador += 1
        progress_bar["value"] = contador
        if JANELA_ABERTA:
            app.update()

    if JANELA_ABERTA:
        limpar_planilha(sheet, wb)
        if desligar_var.get():
            desligar_com_confirmacao()

# ==================================================
# FUNÇÃO DE FECHAMENTO COMPLETO
# ==================================================

def ao_fechar_janela():
    global JANELA_ABERTA
    JANELA_ABERTA = False
    app.destroy()
    sys.exit()

# ==================================================
# INTERFACE GRÁFICA (TKINTER)
# ==================================================

app = tk.Tk()
app.title("Bot Chamados Plantão")
app.geometry("360x310")

app.protocol("WM_DELETE_WINDOW", ao_fechar_janela)

webbrowser.open(SITE_URL)

tk.Button(app, text="Selecionar Imagens", command=selecionar_imagens).pack(pady=5)
tk.Button(app, text="Selecionar Planilha", command=selecionar_planilha, bg="#0d3303", fg="white").pack(pady=5)
tk.Button(app, text="Configurar Dia (F8)", command=configurar_dia, bg="#ffaa00").pack(pady=10)

desligar_var = tk.BooleanVar()
tk.Checkbutton(
    app,
    text="Desligar computador ao finalizar",
    variable=desligar_var
).pack(pady=5)

progress_bar = ttk.Progressbar(app, length=260)
progress_bar.pack(pady=10)

countdown_label = tk.Label(app, text="", fg="red")
countdown_label.pack()

tk.Button(app, text="Iniciar Automação", command=iniciar_automacao, bg="green", fg="white").pack(pady=10)

app.mainloop()