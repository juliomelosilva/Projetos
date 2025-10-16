import openpyxl
import time
import pyautogui
import webbrowser
import tkinter as tk
from tkinter import filedialog, messagebox
import json
import os
import keyboard

# -------------------- CONFIGURAÇÕES INICIAIS --------------------
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.06

config_file = "config_imagens.json"
DEBUG = False  # Mude para True para ver logs adicionais

# -------------------- FUNÇÕES DE CONFIGURAÇÃO --------------------

def salvar_configuracao(imagens):
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(imagens, f, ensure_ascii=False, indent=2)
    print("[i] Configuração salva.")


def carregar_configuracao():
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def configurar_imagens():
    arquivos = filedialog.askopenfilenames(
        title="Selecione as imagens de referência",
        filetypes=[("Imagens", "*.png;*.jpg;*.jpeg")]
    )
    if not arquivos:
        messagebox.showerror("Erro", "Nenhuma imagem selecionada.")
        return

    nomes_campos = [
        "tecnico", "selecao_data", "dia", "empresa", "titulo", "descricao",
        "resolucao", "prioridade", "prioridade_opcao", "categoria", "categoria_opcao", "finalizar"
    ]
    disponiveis = {os.path.basename(a).split('.')[0]: a for a in arquivos}

    imagens = {}
    for nome in nomes_campos:
        if nome in disponiveis:
            imagens[nome] = disponiveis[nome]
        else:
            messagebox.showerror("Erro", f"Falta a imagem '{nome}'. Assegure-se de nomear os arquivos corretamente (sem extensão).")
            return

    salvar_configuracao(imagens)
    messagebox.showinfo("Sucesso", "Imagens configuradas com sucesso.")


def validar_imagens():
    cfg = carregar_configuracao()
    nomes_campos = [
        "tecnico", "selecao_data", "dia", "empresa", "titulo", "descricao",
        "resolucao", "prioridade", "prioridade_opcao", "categoria", "categoria_opcao", "finalizar"
    ]
    faltando = [c for c in nomes_campos if c not in cfg]
    if faltando:
        messagebox.showerror("Erro", f"Imagens faltando: {', '.join(faltando)}")
        return False
    return True

# -------------------- FUNÇÕES DE BUSCA E CLIQUE --------------------

def _clamp_region(x, y, w, h):
    """Garante que a região esteja dentro da tela."""
    sw, sh = pyautogui.size()
    x = max(0, min(x, sw - 1))
    y = max(0, min(y, sh - 1))
    w = max(1, min(w, sw - x))
    h = max(1, min(h, sh - y))
    return int(x), int(y), int(w), int(h)


def localizar_imagem(imagem_path, confiancas=(0.9, 0.85, 0.8, 0.75, 0.7), region=None, tentativa_unica=False):
    """
    Tenta localizar a imagem variando níveis de confiança. Retorna Box center (x,y) se encontrada, senão None.
    Se region for passada, limita a busca a essa área (x, y, w, h).
    """
    if DEBUG:
        print(f"[debug] localizar_imagem: {imagem_path}, region={region}")

    for conf in confiancas:
        try:
            if region:
                pos = pyautogui.locateCenterOnScreen(imagem_path, confidence=conf, region=region)
            else:
                pos = pyautogui.locateCenterOnScreen(imagem_path, confidence=conf)
        except Exception as e:
            if DEBUG:
                print(f"[debug] locate error (conf={conf}): {e}")
            pos = None

        if pos:
            if DEBUG:
                print(f"[debug] encontrado com conf={conf}: {pos}")
            return pos

        if tentativa_unica:
            break

    return None


def clicar_imagem(nome, tentativas=6, intervalo=0.7, region=None):
    """Procura e clica na imagem; retorna a posição se clicou, senão None."""
    cfg = carregar_configuracao()
    if nome not in cfg:
        messagebox.showerror("Erro", f"Imagem '{nome}' não configurada.")
        return None
    imagem = cfg[nome]

    for t in range(tentativas):
        pos = localizar_imagem(imagem, region=region)
        if pos:
            pyautogui.moveTo(pos.x, pos.y, duration=0.12)
            pyautogui.click()
            if DEBUG:
                print(f"[debug] clicado '{nome}' em {pos}")
            return pos
        if DEBUG:
            print(f"[debug] '{nome}' não encontrado. tentativa {t+1}/{tentativas}")
        time.sleep(intervalo)
    return None


def scroll_until_find(nome, max_rolagens=30, passo=-600, confiancas=(0.9, 0.85, 0.8, 0.75, 0.7)):
    """Rola a tela até encontrar a imagem. Retorna posição ou None."""
    cfg = carregar_configuracao()
    if nome not in cfg:
        return None
    imagem = cfg[nome]

    # checa sem rolar primeiro
    pos = localizar_imagem(imagem, confiancas)
    if pos:
        return pos

    for i in range(max_rolagens):
        pyautogui.scroll(passo)
        time.sleep(0.22)
        pos = localizar_imagem(imagem, confiancas)
        if pos:
            if DEBUG:
                print(f"[debug] '{nome}' encontrado apos {i+1} rolagens: {pos}")
            return pos
    return None

# -------------------- FUNÇÃO ESPECIAL PARA 'DIA' --------------------

def tentar_clicar_dia_após_selecao(pos_selecao_data):
    """
    Após clicar em 'selecao_data', tenta localizar e clicar em 'dia' de maneira robusta.
    Estratégia:
    1) Tentar localizar 'dia' na região próxima ao elemento 'selecao_data'.
    2) Se não encontrado, expandir a região e tentar novamente.
    3) Se continuar falhando, tentar alguns offsets (clicks heurísticos) dentro da área do calendário.
    """
    cfg = carregar_configuracao()
    if 'dia' not in cfg:
        print("[!] Imagem 'dia' não configurada.")
        return False

    sw, sh = pyautogui.size()
    # região inicial baseada na posição do botão de seleção de data
    x0 = max(0, pos_selecao_data.x - 220)
    y0 = max(0, pos_selecao_data.y)
    w0 = min(600, sw - x0)
    h0 = min(480, sh - y0)
    region1 = _clamp_region(x0, y0, w0, h0)

    if DEBUG:
        print(f"[debug] tentar_clicar_dia: region1={region1}")

    # 1) tenta na região próxima
    pos_dia = localizar_imagem(cfg['dia'], region=region1)
    if pos_dia:
        pyautogui.moveTo(pos_dia.x, pos_dia.y, duration=0.12)
        pyautogui.click()
        if DEBUG:
            print(f"[debug] clicou dia em {pos_dia} (região 1)")
        return True

    # 2) região expandida (toda metade superior da tela onde calendário pode aparecer)
    x1 = 0
    y1 = max(0, pos_selecao_data.y - 200)
    w1 = sw
    h1 = min(sh - y1, 800)
    region2 = _clamp_region(x1, y1, w1, h1)

    if DEBUG:
        print(f"[debug] tentar_clicar_dia: region2={region2}")

    pos_dia = localizar_imagem(cfg['dia'], region=region2)
    if pos_dia:
        pyautogui.moveTo(pos_dia.x, pos_dia.y, duration=0.12)
        pyautogui.click()
        if DEBUG:
            print(f"[debug] clicou dia em {pos_dia} (região 2)")
        return True

    # 3) heurística de offsets: tenta clicar em algumas posições esperadas do calendário
    offsets = [ (70, 90), (140, 90), (210, 90), (70, 160), (140,160), (210,160) ]
    # baseia-se em pos_selecao_data
    for dx, dy in offsets:
        tx = pos_selecao_data.x + dx
        ty = pos_selecao_data.y + dy
        tx = max(0, min(tx, sw-1))
        ty = max(0, min(ty, sh-1))
        if DEBUG:
            print(f"[debug] tentando offset click em {(tx,ty)}")
        pyautogui.moveTo(tx, ty, duration=0.12)
        pyautogui.click()
        time.sleep(0.18)
        # após clicar, verifica se 'dia' estava neste ponto (procura a imagem no local clicado)
        pos_check = localizar_imagem(cfg['dia'], region=_clamp_region(tx-40, ty-20, 80, 40), confiancas=(0.9,0.85,0.8), tentativa_unica=True)
        if pos_check:
            if DEBUG:
                print(f"[debug] encontrado dia apos offset click: {pos_check}")
            return True

    # 4) última tentativa: busca global com mais tolerância
    pos_dia = localizar_imagem(cfg['dia'], confiancas=(0.8, 0.75, 0.7, 0.65))
    if pos_dia:
        pyautogui.moveTo(pos_dia.x, pos_dia.y, duration=0.12)
        pyautogui.click()
        if DEBUG:
            print(f"[debug] clicou dia em busca global com baixa confiança: {pos_dia}")
        return True

    # tudo falhou
    if DEBUG:
        print("[debug] Não foi possível localizar ou clicar em 'dia'.")
    return False

# -------------------- ROTINA PRINCIPAL --------------------

def iniciar_automacao():
    if not validar_imagens():
        return

    tecnico = entry_tecnico.get().strip()
    if not tecnico:
        messagebox.showerror("Erro", "Informe o nome do técnico.")
        return

    arquivo_excel = filedialog.askopenfilename(title="Selecione o Excel", filetypes=[("Excel", "*.xlsx")])
    if not arquivo_excel:
        return

    messagebox.showinfo("Atenção", "Abra o navegador em tela cheia (F11), zoom 100% (Ctrl+0) e deixe a janela ativa. A automação começará em 3 segundos.")
    time.sleep(3)

    try:
        wb = openpyxl.load_workbook(arquivo_excel)
        sheet = wb['Planilha1'] if 'Planilha1' in wb.sheetnames else wb.active
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir Excel: {e}")
        return

    # abre o site
    webbrowser.open('https://smliveloja.bitrix24.site/plantao/')
    time.sleep(6)

    cfg = carregar_configuracao()

    for linha in sheet.iter_rows(min_row=2, values_only=True):
        Nome, Descricao, Resolucao = linha if len(linha) >= 3 else (linha[0], linha[1] if len(linha) > 1 else None, None)
        if not Nome and not Descricao and not Resolucao:
            print("[i] Linha vazia - finalizando loop.")
            break

        print(f"\n[i] Processando chamado: {Nome}")

        # técnico
        pos = clicar_imagem('tecnico')
        if not pos:
            messagebox.showerror("Erro", "Não foi possível localizar 'tecnico'.")
            return
        keyboard.write(str(tecnico))
        time.sleep(0.28)

        # seleção de data
        pos_data = clicar_imagem('selecao_data')
        if not pos_data:
            messagebox.showerror("Erro", "Não foi possível localizar 'selecao_data'.")
            return
        time.sleep(0.18)

        # tentativa robusta de clicar em 'dia'
        ok_dia = tentar_clicar_dia_após_selecao(pos_data)
        if not ok_dia:
            # tentativa extra: procura globalmente com tolerância
            pos_dia_global = localizar_imagem(cfg['dia'], confiancas=(0.85, 0.8, 0.75, 0.7))
            if pos_dia_global:
                pyautogui.moveTo(pos_dia_global.x, pos_dia_global.y, duration=0.12)
                pyautogui.click()
                ok_dia = True

        if not ok_dia:
            messagebox.showerror("Erro", "Não foi possível selecionar o dia no controle de data. Verifique a imagem de referência 'dia' e o zoom do navegador (Ctrl+0).")
            return

        time.sleep(0.28)

        # empresa
        pos = clicar_imagem('empresa')
        if not pos:
            messagebox.showerror("Erro", "Não foi possível localizar 'empresa'.")
            return
        keyboard.write(str(Nome or ""))
        time.sleep(0.22)

        # título
        pos = clicar_imagem('titulo')
        if not pos:
            messagebox.showerror("Erro", "Não foi possível localizar 'titulo'.")
            return
        keyboard.write(str(Descricao or ""))
        time.sleep(0.22)

        # descrição
        pos = clicar_imagem('descricao')
        if not pos:
            messagebox.showerror("Erro", "Não foi possível localizar 'descricao'.")
            return
        keyboard.write(str(Descricao or ""))
        time.sleep(0.25)

        # rolar até resolução e preencher
        pos_res = scroll_until_find('resolucao', max_rolagens=30, passo=-700)
        if not pos_res:
            # tenta End e nova busca
            pyautogui.press('end')
            time.sleep(0.6)
            pos_res = localizar_imagem(cfg['resolucao'], confiancas=(0.85,0.8,0.75))
        if not pos_res:
            messagebox.showerror("Erro", "Campo 'resolucao' não encontrado.")
            return
        pyautogui.moveTo(pos_res.x, pos_res.y, duration=0.12)
        pyautogui.click()
        keyboard.write(str(Resolucao or ""))
        time.sleep(0.22)
        

        # prioridade: rolar se necessário e clicar
        pos_prior = scroll_until_find('prioridade', max_rolagens=25, passo=-700)
        if not pos_prior:
            # tentativa extra: small scrolls and try
            for _ in range(3):
                pyautogui.scroll(-600)
                time.sleep(0.18)
            pos_prior = localizar_imagem(cfg['prioridade'], confiancas=(0.85,0.8,0.75))
        if not pos_prior:
            messagebox.showerror("Erro", "Campo 'prioridade' não encontrado.")
            return
        pyautogui.moveTo(pos_prior.x, pos_prior.y, duration=0.12)
        pyautogui.click()
        time.sleep(0.18)

        # selecionar opção de prioridade
        pos_prio_op = clicar_imagem('prioridade_opcao', tentativas=8, intervalo=0.4)
        if not pos_prio_op:
            messagebox.showerror("Erro", "Opção de prioridade não encontrada.")
            return
        time.sleep(0.28)

        # categoria: abrir campo
        pos_cat = clicar_imagem('categoria', tentativas=8, intervalo=0.45)
        if not pos_cat:
            messagebox.showerror("Erro", "Campo 'categoria' não encontrado.")
            return
        time.sleep(0.25)

        # rolar até encontrar a opção de categoria (plantão)
        pos_cat_op = scroll_until_find('categoria_opcao', max_rolagens=45, passo=-900)
        if not pos_cat_op:
            # tentar End e nova busca
            pyautogui.press('end')
            time.sleep(0.8)
            pos_cat_op = localizar_imagem(cfg['categoria_opcao'], confiancas=(0.85,0.8,0.75,0.7))

        if not pos_cat_op:
            # tenta imagem alternativa
            if 'categoria_opcao_plantao' in cfg:
                pos_cat_op = scroll_until_find('categoria_opcao_plantao', max_rolagens=20, passo=-800)

        if not pos_cat_op:
            messagebox.showerror("Erro", "Opção de categoria 'plantão' não encontrada.")
            return

        pyautogui.moveTo(pos_cat_op.x, pos_cat_op.y, duration=0.12)
        pyautogui.click()
        time.sleep(0.3)

        # finalizar
        pos_fin = scroll_until_find('finalizar', max_rolagens=25, passo=-800)
        if not pos_fin:
            pyautogui.press('end')
            time.sleep(0.5)
            pos_fin = localizar_imagem(cfg['finalizar'], confiancas=(0.85,0.8))
        if not pos_fin:
            messagebox.showerror("Erro", "Botão 'finalizar' não encontrado.")
            return
        pyautogui.moveTo(pos_fin.x, pos_fin.y, duration=0.12)
        pyautogui.click()

        print(f"[✓] Chamado '{Nome}' enviado.")
        time.sleep(5.2)

    messagebox.showinfo("Concluído", "Processamento finalizado com sucesso.")

# -------------------- INTERFACE SIMPLES --------------------

app = tk.Tk()
app.title("Automação Chamados - Bitrix")
app.geometry("380x300")

lbl = tk.Label(app, text="Nome do Técnico:")
lbl.pack(pady=(10,2))
entry_tecnico = tk.Entry(app)
entry_tecnico.pack(pady=(0,10))

btn_conf = tk.Button(app, text="Configurar Imagens", command=configurar_imagens, bg="#1976D2", fg="white")
btn_conf.pack(pady=8, fill='x', padx=20)

btn_start = tk.Button(app, text="Iniciar Automação", command=iniciar_automacao, bg="#2E7D32", fg="white")
btn_start.pack(pady=8, fill='x', padx=20)

# instrução útil
info = tk.Label(app, text="Antes de iniciar: abra o navegador em tela cheia e use Ctrl+0 (zoom 100%).", fg="#555")
info.pack(pady=(10,6))

app.mainloop()
