import openpyxl
import time
import pyautogui
import webbrowser
import keyboard
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import json
import os
import threading
import sys

# ==================================================
# CONFIGURAÇÕES GERAIS
# ==================================================

SITE_URL = "https://smliveloja.bitrix24.site/plantao/"
CONFIG_DIA = "config_dia.json"

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3

# ==================================================
# VARIÁVEIS GLOBAIS
# ==================================================

IMAGENS = {}
PLANILHA = ""
JANELA_ABERTA = True  # Flag para monitorar o fechamento da interface gráfica

# ==================================================
# FUNÇÃO DE CLIQUE POR IMAGEM
# ==================================================

def clicar_imagem(nome, timeout=15, confidence=0.8):
    inicio = time.time()
    while time.time() - inicio < timeout:
        if not JANELA_ABERTA:
            return
        try:
            pos = pyautogui.locateCenterOnScreen(
                IMAGENS[nome],
                confidence=confidence
            )
            if pos:
                pyautogui.click(pos)
                return
        except Exception:
            pass
        time.sleep(0.4)
    raise Exception(f"Imagem não encontrada: {nome}")

def preencher_campo_seguro(imagem_nome, texto):
    """Clica no campo, limpa dados antigos e fecha sugestões do navegador"""
    if not JANELA_ABERTA: 
        return
    clicar_imagem(imagem_nome)
    time.sleep(0.2)
    
    # Seleciona tudo e apaga (evita acúmulo de texto e limpa ghost prompts)
    keyboard.press_and_release('ctrl+a')
    keyboard.press_and_release('backspace')
    time.sleep(0.1)
    
    # Escreve o conteúdo real da planilha
    keyboard.write(str(texto))
    time.sleep(0.2)
    
    # Pressiona ESC para fechar pop-ups de preenchimento automático do navegador
    keyboard.press_and_release('esc')
    time.sleep(0.2)

# ==================================================
# LIMPAR PLANILHA (APENAS DA LINHA 2 PARA BAIXO)
# ==================================================

def limpar_planilha(sheet, wb):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    wb.save(PLANILHA)

# ==================================================
# DESLIGAMENTO COM CONFIRMAÇÃO + AUTO 10s
# ==================================================

def desligar_com_confirmacao():
    desligar_confirmado = {"cancelado": False, "respondido": False}

    def perguntar():
        resposta = messagebox.askokcancel(
            "Desligamento",
            "Todos os chamados foram registrados com sucesso.\n\n"
            "Desligamento solicitado, deseja cancelar?"
        )
        desligar_confirmado["respondido"] = True
        if resposta:
            os.system("shutdown /s /t 0")
        else:
            desligar_confirmado["cancelado"] = True
            if JANELA_ABERTA:
                countdown_label.config(text="Desligamento cancelado.")

    def contagem():
        for i in range(10, 0, -1):
            if not JANELA_ABERTA or desligar_confirmado["respondido"]:
                return
            countdown_label.config(text=f"Desligando automaticamente em {i}s...")
            time.sleep(1)
        if not desligar_confirmado["cancelado"] and JANELA_ABERTA:
            os.system("shutdown /s /t 0")

    # Threads marcadas como daemon=True para morrerem se a janela principal fechar
    t1 = threading.Thread(target=perguntar, daemon=True)
    t2 = threading.Thread(target=contagem, daemon=True)
    t1.start()
    t2.start()

# ==================================================
# SELEÇÃO DO DIA
# ==================================================

def salvar_dia(coord):
    with open(CONFIG_DIA, "w") as f:
        json.dump(coord, f)

def carregar_dia():
    if os.path.exists(CONFIG_DIA):
        with open(CONFIG_DIA, "r") as f:
            return json.load(f)
    return None

def configurar_dia():
    clicar_imagem("selecao_data.png")
    messagebox.showinfo(
        "Configurar Dia",
        "Posicione o mouse sobre o DIA desejado\n"
        "e pressione F8 para gravar"
    )
    keyboard.wait("f8")
    x, y = pyautogui.position()
    salvar_dia({"x": x, "y": y})
    messagebox.showinfo(
        "Sucesso",
        f"Dia gravado em:\nX={x} | Y={y}"
    )

# ==================================================
# INTERFACE – SELEÇÃO DE ARQUIVOS
# ==================================================

def selecionar_imagens():
    arquivos = filedialog.askopenfilenames(
        title="Selecione todas as imagens",
        filetypes=[("PNG", "*.png")]
    )
    for img in arquivos:
        IMAGENS[os.path.basename(img)] = img
    messagebox.showinfo(
        "Imagens carregadas",
        f"{len(IMAGENS)} imagens carregadas"
    )

def selecionar_planilha():
    global PLANILHA
    PLANILHA = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Excel", "*.xlsx")]
    )

# ==================================================
# AUTOMAÇÃO PRINCIPAL
# ==================================================

def iniciar_automacao():
    nome_tecnico = nome_entry.get()

    if not nome_tecnico:
        messagebox.showerror("Erro", "Informe o nome do técnico")
        return
    if not PLANILHA:
        messagebox.showerror("Erro", "Selecione a planilha")
        return

    dia = carregar_dia()
    if not dia:
        messagebox.showerror("Erro", "Configure o dia antes de iniciar")
        return

    wb = openpyxl.load_workbook(PLANILHA)
    sheet = wb.active

    total = sum(1 for row in sheet.iter_rows(min_row=2) if row[0].value)
    progress_bar["maximum"] = total
    progress_bar["value"] = 0
    contador = 0

    for linha in sheet.iter_rows(min_row=2):
        if not JANELA_ABERTA:  # Interrompe o loop imediatamente se a janela gráfica sumir
            break

        empresa = linha[0].value
        titulo = linha[1].value
        descricao = linha[1].value
        resolucao = linha[2].value

        if not empresa:
            break

        # 1. Nome do Técnico
        preencher_campo_seguro("tecnico.png", nome_tecnico)

        # 2. Seleção de Data (Foco controlado por cliques + ESC)
        clicar_imagem("selecao_data.png")
        time.sleep(0.5)
        pyautogui.click(dia["x"], dia["y"])
        time.sleep(0.5)
        keyboard.press_and_release('esc')
        time.sleep(0.4)

        # 3. Empresa
        preencher_campo_seguro("empresa.png", empresa)

        # 4. Título
        preencher_campo_seguro("titulo.png", titulo)

        # 5. Descrição
        preencher_campo_seguro("descricao.png", descricao)

        # 6. Resolução
        preencher_campo_seguro("resolucao.png", resolucao)

        # Desce a tela do formulário para visualizar os blocos dinâmicos finais
        pyautogui.scroll(-2000)
        time.sleep(0.5)

        # 7. Prioridade
        clicar_imagem("prioridade.png")
        time.sleep(0.5)
        clicar_imagem("prioridade_opcao.png")
        time.sleep(0.5)

        # 8. Categoria - Movimento por Tempo com Repetição Contínua
        clicar_imagem("categoria.png")
        time.sleep(0.6)  # Tempo para o drop-down processar a abertura
        
        # Move o ponteiro 100 pixels para baixo para focar na área interna da lista
        pyautogui.moveRel(0, 100)
        time.sleep(0.2)

        if JANELA_ABERTA:
            # Marca o momento que iniciou a contagem
            tempo_inicio = time.time()
            
            # Executa o pressionamento repetido por exatamente 4 segundos
            while time.time() - tempo_inicio < 4.0:
                if not JANELA_ABERTA:
                    break
                keyboard.press_and_release('down')
                time.sleep(0.05) # Envia o sinal várias vezes por segundo para descer a lista
            
            # Aguarda 1 segundo após o término dos 4 segundos
            time.sleep(1.0)
            
            # Confirma a seleção com o Enter
            keyboard.press_and_release('enter')
            time.sleep(0.5)

        # 9. Finalizar Registro
        clicar_imagem("finalizar.png")
        time.sleep(6)

        contador += 1
        progress_bar["value"] = contador
        if JANELA_ABERTA:
            app.update()

    if JANELA_ABERTA:
        limpar_planilha(sheet, wb)
        if desligar_var.get():
            desligar_com_confirmacao()

# ==================================================
# FUNÇÃO DE FECHAMENTO COMPLETO
# ==================================================

def ao_fechar_janela():
    """Para imediatamente todas as rotinas em background e mata o processo Python"""
    global JANELA_ABERTA
    JANELA_ABERTA = False
    app.destroy()
    sys.exit()

# ==================================================
# INTERFACE GRÁFICA (TKINTER)
# ==================================================

app = tk.Tk()
app.title("Bot Chamados Plantão")
app.geometry("360x360")

# Vincula o botão fechar 'X' do Windows à nossa rotina de encerramento total
app.protocol("WM_DELETE_WINDOW", ao_fechar_janela)

webbrowser.open(SITE_URL)

tk.Label(app, text="Nome do Técnico").pack(pady=5)
nome_entry = tk.Entry(app, width=30)
nome_entry.pack()

tk.Button(app, text="Selecionar Imagens", command=selecionar_imagens).pack(pady=5)
tk.Button(app, text="Selecionar Planilha", command=selecionar_planilha, bg="#0d3303", fg="white").pack(pady=5)
tk.Button(app, text="Configurar Dia (F8)", command=configurar_dia, bg="#ffaa00").pack(pady=10)

desligar_var = tk.BooleanVar()
tk.Checkbutton(
    app,
    text="Desligar computador ao finalizar",
    variable=desligar_var
).pack(pady=5)

progress_bar = ttk.Progressbar(app, length=260)
progress_bar.pack(pady=10)

countdown_label = tk.Label(app, text="", fg="red")
countdown_label.pack()

tk.Button(app, text="Iniciar Automação", command=iniciar_automacao, bg="green", fg="white").pack(pady=10)

app.mainloop()